"""Traceable Excel workbook export."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from direct_infusion_quant.calibration import CalibrationResult
from direct_infusion_quant.export.tables import result_tables
from direct_infusion_quant.models import AnalysisProject
from direct_infusion_quant.processing import FileProcessingResult


def export_excel_workbook(
    path: Path,
    project: AnalysisProject,
    results: dict[str, FileProcessingResult],
    calibration: CalibrationResult | None,
) -> None:
    """Write results and the complete settings summary using openpyxl."""

    destination = path.expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        for name, rows in result_tables(project, results, calibration).items():
            pd.DataFrame(rows).to_excel(writer, sheet_name=name[:31], index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
                cell.alignment = Alignment(vertical="top")
            if sheet.max_row > 1:
                sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                lengths = [len(str(cell.value or "")) for cell in column]
                width = min(max(lengths) + 2, 60)
                sheet.column_dimensions[column[0].column_letter].width = width
                for cell, length in zip(column, lengths, strict=True):
                    if length <= 58:
                        continue
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    wrapped_height = min(15 * ((length // 58) + 1), 60)
                    current_height = sheet.row_dimensions[cell.row].height or 15
                    sheet.row_dimensions[cell.row].height = max(
                        current_height, wrapped_height
                    )
