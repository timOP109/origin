"""Packaged executable diagnostic for core Windows deployment workflows."""

from __future__ import annotations

import sys
from datetime import UTC, datetime
from pathlib import Path

from matplotlib.figure import Figure
from openpyxl import load_workbook

from direct_infusion_quant.export import export_excel_workbook, export_png_plots
from direct_infusion_quant.gui.workers import capture_source_provenance
from direct_infusion_quant.io import MzMLBackend, create_mzml_reader
from direct_infusion_quant.models import (
    AnalysisProject,
    AnalyteTarget,
    BlankCorrectionMethod,
    CalibrationSettings,
    ExtractionWindow,
    ProcessingSettings,
    QuantifierMode,
    SampleRecord,
    SampleType,
    SummaryMethod,
    ToleranceUnit,
)
from direct_infusion_quant.persistence import load_project, save_project
from direct_infusion_quant.processing import WarningThresholds, process_file


def run_from_arguments(arguments: list[str]) -> int:
    """Run import/process/plot/project/Excel checks with explicit settings."""

    if len(arguments) not in {6, 7}:
        print(
            "Usage: --packaging-smoke SOURCE_MZML OUTPUT_DIR START_S END_S "
            "TARGET_MZ TOLERANCE_DA [pymzml|pyopenms]",
            file=sys.stderr,
        )
        return 2
    source = Path(arguments[0]).resolve()
    destination = Path(arguments[1]).resolve()
    start, end, target, tolerance = map(float, arguments[2:6])
    backend = MzMLBackend(arguments[6]) if len(arguments) == 7 else MzMLBackend.PYMZML
    destination.mkdir(parents=True, exist_ok=True)

    window = ExtractionWindow(
        name="Packaging smoke quantifier",
        target_mz=target,
        tolerance=tolerance,
        tolerance_unit=ToleranceUnit.DA,
    )
    analyte = AnalyteTarget(
        name="Packaging smoke analyte",
        windows=[window],
        quantifier_mode=QuantifierMode.SINGLE,
        quantifier_window_ids=[window.id],
    )
    sample = SampleRecord(
        path=source,
        sample_name=source.stem,
        sample_type=SampleType.UNKNOWN,
        source_provenance=capture_source_provenance(source, backend=backend),
    )
    settings = ProcessingSettings(
        mzml_backend=backend,
        ms_level=1,
        time_start_seconds=start,
        time_end_seconds=end,
        summary_method=SummaryMethod.MEDIAN,
    )
    processed_at = datetime.now(UTC)
    project = AnalysisProject(
        name="Packaged application smoke test",
        samples=[sample],
        analytes=[analyte],
        active_analyte_id=analyte.id,
        processing=settings,
        calibration=CalibrationSettings(blank_correction=BlankCorrectionMethod.NONE),
        last_processing_timestamp_utc=processed_at,
    )
    result = process_file(
        create_mzml_reader(backend).iter_spectra(source),
        analyte.windows,
        settings,
        WarningThresholds(),
        quantifier_window_id=window.id,
    )
    results = {str(sample.id): result}

    project_path = destination / "packaging-smoke.diq.json"
    workbook_path = destination / "packaging-smoke.xlsx"
    save_project(project, project_path)
    reopened = load_project(project_path)
    export_excel_workbook(workbook_path, reopened, results, None)

    figure = Figure(figsize=(7, 4))
    axes = figure.add_subplot(111)
    axes.plot(
        [scan.elapsed_time_seconds for scan in result.scans],
        [scan.window_responses[window.id] for scan in result.scans],
    )
    axes.set_xlabel("Elapsed acquisition time (s)")
    axes.set_ylabel("Summed intensity")
    export_png_plots(destination, {"packaging-smoke-spray-response": figure})

    workbook = load_workbook(workbook_path, read_only=True)
    required_sheets = {"Samples", "Processing Settings", "Scan Summary"}
    if not required_sheets.issubset(workbook.sheetnames):
        raise RuntimeError("Packaged Excel export is missing required sheets.")
    print(
        "PACKAGING_SMOKE_OK "
        f"backend={backend.value} scans={len(result.scans)} "
        f"response={result.quantification_response:g} "
        f"project={project_path} workbook={workbook_path}"
    )
    return 0
