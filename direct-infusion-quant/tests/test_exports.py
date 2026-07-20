"""Tests for reproducibility-focused CSV, Excel, and PNG exports."""

from datetime import UTC, datetime
from pathlib import Path

import numpy as np
import openpyxl
import pytest
from matplotlib.figure import Figure

from direct_infusion_quant.calibration import CalibrationInput, calibrate_and_quantify
from direct_infusion_quant.export import (
    export_csv_tables,
    export_excel_workbook,
    export_png_plots,
)
from direct_infusion_quant.export.tables import CSV_NAMES, SHEET_NAMES, result_tables
from direct_infusion_quant.io.base import SpectrumRecord
from direct_infusion_quant.models import (
    AnalysisProject,
    AnalyteTarget,
    CalibrationSettings,
    ExtractionWindow,
    MzMLBackend,
    ProcessingSettings,
    QuantifierMode,
    RegressionModel,
    SampleRecord,
    SampleType,
    SourceFileProvenance,
    StabilityAssessmentRecord,
    StabilityCandidateRecord,
    StabilityTraceMode,
    ToleranceUnit,
)
from direct_infusion_quant.processing import WarningThresholds, process_file


def project_and_results(tmp_path: Path):
    window = ExtractionWindow(
        name="quantifier",
        target_mz=500,
        tolerance=0.1,
        tolerance_unit=ToleranceUnit.DA,
        charge=2,
    )
    analyte = AnalyteTarget(
        name="peptide",
        molecular_weight=1000,
        notes="review note",
        windows=[window],
        quantifier_mode=QuantifierMode.SINGLE,
        quantifier_window_ids=[window.id],
    )
    source_path = tmp_path / "external.mzML"
    source_path.write_bytes(b"source")
    provenance = SourceFileProvenance(
        file_size_bytes=6,
        modified_time_ns=source_path.stat().st_mtime_ns,
        modified_time_utc=datetime.fromtimestamp(source_path.stat().st_mtime, UTC),
        sha256="41cf6794ba4200b839c1322c8af1f7eea0e78e94b1f7b156d0af752dc1344e6c",
        spectrum_count=1,
        ms_levels=[1],
        is_centroided=True,
        captured_at_utc=datetime(2026, 7, 18, tzinfo=UTC),
    )
    sample = SampleRecord(
        path=source_path,
        sample_name="unknown",
        sample_type=SampleType.UNKNOWN,
        dilution_factor=2,
        source_provenance=provenance,
        time_start_seconds=5.0,
        stability_assessment=StabilityAssessmentRecord(
            assessed_at_utc=datetime(2026, 7, 18, 11, tzinfo=UTC),
            trace_mode=StabilityTraceMode.TIC,
            candidates=[
                StabilityCandidateRecord(
                    rank=1,
                    start_seconds=5,
                    end_seconds=6,
                    scan_count=10,
                    trace_median_response=100,
                    trace_robust_cv_percent=2,
                    trace_relative_drift_percent=1,
                    trace_zero_fraction=0,
                    trace_signal_fraction=1,
                    score=3,
                    analyte_median_response=42,
                    analyte_robust_cv_percent=4,
                    analyte_relative_drift_percent=2,
                    analyte_zero_fraction=0,
                )
            ],
        ),
    )
    settings = ProcessingSettings(
        mzml_backend=MzMLBackend.PYOPENMS,
        time_start_seconds=0,
        time_end_seconds=1,
    )
    project = AnalysisProject(
        name="export",
        samples=[sample],
        analytes=[analyte],
        active_analyte_id=analyte.id,
        processing=settings,
        calibration=CalibrationSettings(regression_model=RegressionModel.QUADRATIC),
        last_processing_timestamp_utc=datetime(2026, 7, 18, 12, tzinfo=UTC),
    )
    spectra = [
        SpectrumRecord(
            native_id="scan=1",
            index=1,
            ms_level=1,
            elapsed_time_seconds=0.5,
            mz=np.asarray([500.0]),
            intensity=np.asarray([42.0]),
        )
    ]
    result = process_file(
        spectra,
        [window],
        settings,
        WarningThresholds(),
        quantifier_window_id=window.id,
    )
    return project, {str(sample.id): result}


def test_csv_and_excel_contain_all_reproducibility_tables(tmp_path: Path) -> None:
    project, results = project_and_results(tmp_path)
    csv_paths = export_csv_tables(tmp_path / "csv", project, results, None)
    assert {path.name for path in csv_paths} == set(CSV_NAMES.values())
    scan_text = (tmp_path / "csv" / "scan_summary.csv").read_text(encoding="utf-8")
    assert "elapsed_acquisition_time_seconds" in scan_text
    assert "42.0" in scan_text
    samples_text = (tmp_path / "csv" / "samples.csv").read_text(encoding="utf-8")
    assert "source_file_sha256" in samples_text
    assert "41cf6794" in samples_text
    settings_text = (tmp_path / "csv" / "processing_settings.csv").read_text(
        encoding="utf-8"
    )
    assert "selected_ms_level" in settings_text
    assert "mzml_reader_backend" in settings_text
    assert "pyopenms" in settings_text
    assert "pooled_median" in settings_text
    assert "quadratic" in settings_text
    assert "stability_trace_mode" in settings_text
    assessment_text = (tmp_path / "csv" / "stability_assessment.csv").read_text(
        encoding="utf-8"
    )
    assert "trace_robust_cv_percent" in assessment_text
    assert "analyte_robust_cv_percent" in assessment_text

    workbook_path = tmp_path / "analysis.xlsx"
    export_excel_workbook(workbook_path, project, results, None)
    workbook = openpyxl.load_workbook(workbook_path)
    assert tuple(workbook.sheetnames) == SHEET_NAMES
    sample_headers = next(
        workbook["Samples"].iter_rows(min_row=1, max_row=1, values_only=True)
    )
    assert "source_file_size_bytes" in sample_headers
    assert "source_file_sha256" in sample_headers
    assert workbook["Samples"][1][0].font.bold is True
    assert workbook["Samples"][1][0].fill.fgColor.rgb == "001F4E78"
    software_rows = list(workbook["Software Versions"].iter_rows(values_only=True))
    assert any(row[0] == "DirectInfusionQuant" for row in software_rows)
    assert any("reproducibility and review" in str(row[1]) for row in software_rows)


def test_png_plot_export(tmp_path: Path) -> None:
    figure = Figure()
    axes = figure.add_subplot(111)
    axes.plot([0, 1], [1, 2])
    paths = export_png_plots(tmp_path, {"spray-response": figure})
    assert paths == [tmp_path.resolve() / "spray-response.png"]
    assert paths[0].read_bytes().startswith(b"\x89PNG")


def test_polynomial_coefficients_are_exported() -> None:
    samples = [
        SampleRecord(
            path=Path(f"s{x}.mzML"),
            sample_name=f"s{x}",
            sample_type=SampleType.STANDARD,
            concentration=float(x),
            concentration_unit="mg/mL",
        )
        for x in range(5)
    ]
    settings = CalibrationSettings(
        blank_correction="none", regression_model=RegressionModel.QUADRATIC
    )
    calibration = calibrate_and_quantify(
        [
            CalibrationInput(
                sample_id=str(sample.id),
                sample_type=sample.sample_type,
                concentration=sample.concentration,
                concentration_unit=sample.concentration_unit,
                file_response=1
                + 2 * sample.concentration
                + 3 * sample.concentration**2,
            )
            for sample in samples
        ],
        settings,
    )
    project = AnalysisProject(
        name="polynomial export", samples=samples, calibration=settings
    )
    tables = result_tables(project, {}, calibration)
    statistics = tables["Calibration Statistics"][0]
    assert statistics["regression_model"] == "quadratic"
    assert statistics["polynomial_order"] == 2
    assert statistics["coefficient_c0"] == pytest.approx(1)
    assert statistics["coefficient_c1"] == pytest.approx(2)
    assert statistics["coefficient_c2"] == pytest.approx(3)
